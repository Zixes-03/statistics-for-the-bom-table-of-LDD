import numpy as np
import openpyxl as pyxl
import io
import os
import re
import gc
from PIL import Image
import datetime
import time
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog

# 表格图片dict存储
def retrieve_image_dict(ws):
    image_array_dict = {}
    for image in ws._images:
        image_row = image.anchor._from.row
        brick_code = ws.cell(row = image_row + 1, column=1).value
        brick_id = ws.cell(row = image_row + 1, column=4).value
        brick_color = ws.cell(row = image_row + 1, column=5).value
        if brick_code == None:
            image_array_dict[f'{brick_id} + {brick_color}'] = np.array(Image.open(image.ref).convert("RGBA"))
        else:
            image_array_dict[f'{brick_code}'] = np.array(Image.open(image.ref).convert("RGBA"))
    return image_array_dict

# 重写Image方法
class rewrite_Image(pyxl.drawing.image.Image):
    def _data(self):
        img = pyxl.drawing.image._import_image(self.ref)
        fp = io.BytesIO()
        img.save(fp, format="png")
        fp.seek(0)

        return fp.read()

# 插入图片
def insert_image(ws_import, image_dict):
    for i in range(2,ws_import.max_row):
        brick_code = ws_import.cell(row=i,column=1).value
        brick_id = ws_import.cell(row=i,column=4).value
        brick_color = ws_import.cell(row=i,column=5).value
        if brick_code == None:
            image_insert = Image.fromarray(image_dict[f'{brick_id} + {brick_color}'])
        else:
            image_insert = Image.fromarray(image_dict[f'{brick_code}'])
        image_insert = rewrite_Image(image_insert)
        image_insert.width, image_insert.height = image_insert.width * 0.57, image_insert.height * 0.67
        ws_import.add_image(image_insert, anchor='C' + str(i))

# 更新进度条
def update_progress(value):
    progress_bar['value'] = value
    root.update()

# 统计缺少的零件
def insufficient_statistics(owned_path, new_path, zero_judge = 0):
    update_progress(0)
    wb_owned = pyxl.load_workbook(owned_path)
    ws_owned = wb_owned.active
    wb_new = pyxl.load_workbook(new_path)
    ws_new = wb_new.active
    for row_num in range(2, ws_new.max_row):
        ws_new.row_dimensions[row_num].height = 14.25
    # 删除图片
#     ws_owned_image_array_dict = retrieve_image_dict(ws_owned)
    for i in range(len(ws_owned._images)-1, -1, -1):
        del ws_owned._images[i]
    ws_new_image_array_dict = retrieve_image_dict(ws_new)
    for i in range(len(ws_new._images)-1, -1, -1):
        del ws_new._images[i]
#     ws_owned_image_array_dict.update(ws_new_image_array_dict)
    update_progress(25)
    # 统计缺少的零件
    ws_new.delete_cols(7)
    ws_new.cell(row=1,column=7).value = 'add_QTY'
    for i in range(ws_new.max_row - 1, 1, -1):
        brick_code_new = ws_new.cell(row=i,column=1).value
        brick_id_new = ws_new.cell(row=i,column=4).value
        brick_color_new = re.findall('\- (.+)',ws_new.cell(row=i,column=5).value)[0]
    #     brick_color_new = ws_new.cell(row=i,column=5).value
        brick_qty_new = ws_new.cell(row=i,column=6).value
        for j in range(2,ws_owned.max_row):
            brick_code_owned = ws_owned.cell(row=j,column=1).value
            brick_id_owned = ws_owned.cell(row=j,column=4).value
            brick_color_owned = re.findall('\- (.+)',ws_owned.cell(row=j,column=5).value)[0]
    #         brick_color_owned = ws_owned.cell(row=j,column=5).value
            brick_qty_owned = ws_owned.cell(row=j,column=6).value
            if brick_id_new == brick_id_owned and brick_color_new == brick_color_owned and brick_code_new == None:
                if brick_qty_owned <= brick_qty_new:
                    ws_new.cell(row=i,column=7).value = brick_qty_new - brick_qty_owned
                    break
                else:
                    ws_new.delete_rows(i)
                    break
            elif brick_code_owned != None and brick_code_new == brick_code_owned:
                if brick_qty_owned <= brick_qty_new:
                    ws_new.cell(row=i,column=7).value = brick_qty_new - brick_qty_owned
                    break
                else:
                    ws_new.delete_rows(i)
                    break
        if ws_new.cell(row=i,column=7).value == None:
            ws_new.cell(row=i,column=7).value = brick_qty_new
    update_progress(50)
    # 删除添加数小于等于0的零件
    if zero_judge == 0:
        for i in range(ws_new.max_row-1, 1, -1):
            if ws_new.cell(row=i ,column=7).value <= 0:
                ws_new.delete_rows(i)
    # 统计零件总数
    ws_new.cell(row = ws_new.max_row, column=6).value = f'=sum(F2:F{ws_new.max_row - 1})'
    ws_new.cell(row = ws_new.max_row, column=7).value = f'=sum(G2:G{ws_new.max_row - 1})'
    # 修改单元格格式
    for cols in ws_new.iter_cols():
        for cell in cols:
            ws_new[cell.coordinate].font = pyxl.styles.Font(name = '等线')
            ws_new[cell.coordinate].alignment = pyxl.styles.Alignment(vertical='center')
    for row_num in range(2, ws_new.max_row):
        ws_new.row_dimensions[row_num].height = 66
    # 插入图片
    insert_image(ws_new, ws_new_image_array_dict)
    update_progress(75)
    # 导出文件
    time_now = datetime.datetime.now().strftime('%Y-%m-%d_%H.%M.%S')
    new_path_folder = os.path.dirname(os.path.abspath(brick_new_path.get()))
    purchase_list_path = f'{new_path_folder}\\add_brick_{time_now}.xlsx'
    wb_new.save(purchase_list_path)
    # 释放内存
    del wb_owned, ws_owned, wb_new, ws_new, ws_new_image_array_dict
    gc.collect()
    update_progress(100)

# 汇总零件
def brick_summarize(owned_path, purchased_path, zero_judge=1):
    update_progress(0)
    wb_owned = pyxl.load_workbook(owned_path)
    ws_owned = wb_owned.active
    wb_purchased = pyxl.load_workbook(purchased_path)
    ws_purchased = wb_purchased.active
    # 删除图片
    ws_owned_image_array_dict = retrieve_image_dict(ws_owned)
    for i in range(len(ws_owned._images)-1, -1, -1):
        del ws_owned._images[i]
    ws_purchased_image_array_dict = retrieve_image_dict(ws_purchased)
    for i in range(len(ws_purchased._images)-1, -1, -1):
        del ws_purchased._images[i]
    ws_owned_image_array_dict.update(ws_purchased_image_array_dict)
    update_progress(25)
    # 汇总新拥有的零件
    for i in range(2, ws_purchased.max_row):
        brick_code_purchased = ws_purchased.cell(row=i,column=1).value
        brick_id_purchased = ws_purchased.cell(row=i,column=4).value
        brick_color_purchased = re.findall('\- (.+)',ws_purchased.cell(row=i,column=5).value)[0]
    #     brick_color_new = ws_purchased.cell(row=i,column=5).value
        brick_qty_purchased = ws_purchased.cell(row=i,column=7).value
        judge = 0
        for j in range(2,ws_owned.max_row):
            judge = judge + 1
            brick_code_owned = ws_owned.cell(row=j,column=1).value
            brick_id_owned = ws_owned.cell(row=j,column=4).value
            brick_color_owned = re.findall('\- (.+)',ws_owned.cell(row=j,column=5).value)[0]
    #         brick_color_owned = ws_owned.cell(row=j,column=5).value
            brick_qty_owned = ws_owned.cell(row=j,column=6).value
            if brick_id_purchased == brick_id_owned and brick_color_purchased == brick_color_owned and brick_code_owned == None:
                ws_owned.cell(row=j,column=6).value = brick_qty_purchased + brick_qty_owned
                break
            elif brick_code_owned != None and brick_code_owned == brick_code_purchased:
                ws_owned.cell(row=j,column=6).value = brick_qty_purchased + brick_qty_owned
                break
        if judge == ws_owned.max_row-2:
            ws_owned.insert_rows(ws_owned.max_row)
            for col_num in range(1,6):
                ws_owned.cell(row=ws_owned.max_row-1,column=col_num).value = ws_purchased.cell(row=i,column=col_num).value
            ws_owned.cell(row=ws_owned.max_row-1,column=6).value = brick_qty_purchased
    update_progress(50)
    # 删除数量小于等于0的零件
    for row_num in range(2, ws_owned.max_row):
        ws_owned.row_dimensions[row_num].height = 14.25
    if zero_judge == 1:
        for i in range(ws_owned.max_row-1, 1, -1):
            if ws_owned.cell(row=i ,column=6).value <= 0:
                ws_owned.delete_rows(i)
    # 统计零件总数
    ws_owned.cell(row = ws_owned.max_row, column=6).value = f'=sum(F2:F{ws_owned.max_row - 1})'
    # 修改单元格格式
    for cols in ws_owned.iter_cols():
        for cell in cols:
            ws_owned[cell.coordinate].font = pyxl.styles.Font(name = '等线')
            ws_owned[cell.coordinate].alignment = pyxl.styles.Alignment(vertical='center')
    for row_num in range(2, ws_owned.max_row):
        ws_owned.row_dimensions[row_num].height = 66
    # 插入图片
    insert_image(ws_owned, ws_owned_image_array_dict)
    update_progress(75)
    # 导出文件
    time_now = datetime.datetime.now().strftime('%Y-%m-%d_%H.%M.%S')
    owned_path_folder = os.path.dirname(os.path.abspath(brick_owned_path.get()))
    owned_list_path = f'{owned_path_folder}\\brick_list_{time_now}.xlsx'
    wb_owned.save(owned_list_path)
    # 释放内存
    del wb_owned, ws_owned, wb_purchased, ws_purchased, ws_purchased_image_array_dict, ws_owned_image_array_dict
    gc.collect()
    update_progress(100)

def buttom_function(buttom_num):
    if buttom_num == 'buttom_01':
        selected_file_path = filedialog.askopenfilename(filetypes=[('xlsx', '*.xlsx')])
        brick_owned_path.set(selected_file_path)
    elif buttom_num == 'buttom_02':
        selected_file_path = filedialog.askopenfilename(filetypes=[('xlsx', '*.xlsx')])
        brick_new_path.set(selected_file_path)
    elif buttom_num == 'buttom_03':
        selected_file_path = filedialog.askopenfilename(filetypes=[('xlsx', '*.xlsx')])
        brick_purchased_path.set(selected_file_path)
    elif buttom_num == 'buttom_04':
        insufficient_statistics(brick_owned_path.get(),brick_new_path.get(),zero_judge=check_value_01.get())
    elif buttom_num == 'buttom_05':
        brick_summarize(brick_owned_path.get(), brick_purchased_path.get(),zero_judge=check_value_02.get())

def judge_buttom():
    if len(brick_owned_path.get()) != 0 and len(brick_new_path.get()) != 0:
        buttom_04['state'] = tk.NORMAL
    else:
        buttom_04['state'] = tk.DISABLED
    if len(brick_owned_path.get()) != 0 and len(brick_purchased_path.get()) != 0:
        buttom_05['state'] = tk.NORMAL
    else:
        buttom_05['state'] = tk.DISABLED
    root.after(100,judge_buttom)

root = tk.Tk()
root.minsize(670,180)
root.maxsize(670,180)
root.title("LEGO Brick Statistics")
# 第一个按钮
brick_owned_path = tk.StringVar()
tk.Label(root, text="文件路径：").grid(row=1, column=0, padx=10, pady=5)
tk.Entry(root, textvariable = brick_owned_path).grid(row=1,column=1,columnspan=3,ipadx=150)
buttom_01 = tk.Button(root, text="选择已拥有的零件表", 
                      command=lambda buttom_num = 'buttom_01' : buttom_function(buttom_num))
buttom_01.grid(row=1, column=4, padx=10)
# 第二个按钮
brick_new_path = tk.StringVar()
tk.Label(root, text="文件路径：").grid(row=2, column=0, padx=10, pady=5)
tk.Entry(root, textvariable = brick_new_path).grid(row=2,column=1,columnspan=3,ipadx=150)
buttom_02 = tk.Button(root, text="选择新导出的零件表", 
                      command=lambda buttom_num = 'buttom_02' : buttom_function(buttom_num))
buttom_02.grid(row=2, column=4, padx=10)
# 第三个按钮
brick_purchased_path = tk.StringVar()
tk.Label(root, text="文件路径：").grid(row=3, column=0, padx=10, pady=5)
tk.Entry(root, textvariable = brick_purchased_path).grid(row=3,column=1,columnspan=3,ipadx=150)
buttom_03 = tk.Button(root, text="选择新添加的零件表", 
                      command=lambda buttom_num = 'buttom_03' : buttom_function(buttom_num))
buttom_03.grid(row=3, column=4, padx=10)
# 第四个按钮
buttom_04 = tk.Button(root, text="统计并导出需要添加的零件", 
                      command=lambda buttom_num = 'buttom_04' : buttom_function(buttom_num), state = 'disable')
buttom_04.grid(row=4, column=1, padx=10, pady=5)
# 第五个按钮
buttom_05 = tk.Button(root, text="合并添加的零件并导出表格", 
                      command=lambda buttom_num = 'buttom_05' : buttom_function(buttom_num), state = 'disable')
buttom_05.grid(row=4, column=3, padx=10, pady=5)
# 可选项目
check_value_01 = tk.IntVar()
check_value_01.set(1)
check_01 = tk.Checkbutton(root, text="统计数量恰好\n的零件", variable = check_value_01, onvalue=1, offvalue=0)
check_01.grid(row=4, column=2)
check_value_02 = tk.IntVar()
check_value_02.set(1)
check_02 = tk.Checkbutton(root, text="删除数量为0\n的零件", variable = check_value_02, onvalue=1, offvalue=0)
check_02.grid(row=4, column=4)
# 进度条
progress_bar = ttk.Progressbar(root, length=100, mode="determinate", orient="horizontal")
progress_bar.grid(row=5, column=0, columnspan=5, ipadx=260, pady=5)

# 持续判断按钮状态
root.after(100,judge_buttom)
root.mainloop()